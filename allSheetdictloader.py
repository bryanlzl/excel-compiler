import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
import glob
from openpyxl.utils.dataframe import dataframe_to_rows

## get dictionary of all sheets in all workbooks ##
def getsheetdict():
    sheetdict = dict()
    for xlsxfile in glob.glob("./To be compiled/*.xlsx"):
        wb = load_workbook(xlsxfile)
        sheetdict[xlsxfile.replace('./To be compiled\\', '')] = []
        for i in wb.worksheets:
            if i.sheet_state == "visible":
                sheetdict[xlsxfile.replace('./To be compiled\\', '')].append(i.title)
    return sheetdict
