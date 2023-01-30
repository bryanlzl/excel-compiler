import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
import glob
from openpyxl.utils.dataframe import dataframe_to_rows

def indexmatcher(compileddf, iminstructdictc):
    for file in iminstructdictc:
        ## Loading trading worksheet to create dataframe of trading list ##
        imwb = load_workbook("./Index match/"+f'{file}', data_only=True)
        imws = imwb.active
        imdf = pd.DataFrame(imws.values)
        imdf.columns = imdf.iloc[0]
        imdf = imdf[1:]
        imdict = dict(zip(imdf[iminstructdictc[file][0]], imdf[iminstructdictc[file][1]]))
        compileddf[iminstructdictc[file][3]] = compileddf[iminstructdictc[file][2]].apply(lambda x: imdict.get(x)).fillna('')
    return compileddf