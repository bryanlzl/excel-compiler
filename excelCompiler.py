import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import glob
import os
import allSheetdictloader as asdl
import indexMatch as im
import excelExporter as ee


def getimsheetdict():
    imsheetdict = dict()
    for xlsxfile in glob.glob("./Index match/*.xlsx"):
        wb = load_workbook(xlsxfile)
        imsheetdict[xlsxfile.replace('./Index match\\','').replace('.xlsx','')] = []
        for i in wb.active[1]:
            imsheetdict[xlsxfile.replace('./Index match\\','').replace('.xlsx','')].append(i.value)
    return imsheetdict

def checkindexmatch(compileddf, iminstructdict):
    ## determine if the sets of index matching instructions are valid with the excel files on hand ##
    iminstructdictc = {} # verified dict of index matching instructions
    imsheetdict = getimsheetdict()
    for file in iminstructdict:
            # check if index match (instruc) and target index columns exist, and 
        if (iminstructdict[file][0] in imsheetdict[file.replace('.xlsx','')]) and (iminstructdict[file][1] in imsheetdict[file.replace('.xlsx','')]):
                    # check if index match (dict) and destination column
            if (iminstructdict[file][2] in compileddf.columns) and (iminstructdict[file][3] in compileddf.columns):
                iminstructdictc[file] = iminstructdict[file]
    return iminstructdictc 

## Loading template worksheet ##
def gettemplatesheet(ctemplatepath, formattype):
    templatebook = load_workbook(ctemplatepath, data_only=True)
    templatesheet = templatebook[formattype]
    return templatesheet

## Loading dataframe for current compilation task ##
def getrawdf(xlsxfile, cwstype):
    wb = load_workbook(xlsxfile)
    ws = wb[cwstype]
    ## converting excel ws to dataframe ##
    wsdf = pd.DataFrame(ws.values)
    wsdf.columns = wsdf.iloc[0]  
    wsdf = wsdf[1:]  
    return wsdf

## Loading the correct format for current compilation task ##
def getcorrectformat(templatesheet, cwstype, formattype):
    correctformat = []
    for worksheetname in templatesheet:
        for row in range(1, templatesheet.max_row+1):
            if cwstype == templatesheet.cell(row=row, column=1).value:
                for col in range(1, templatesheet.max_column+1):
                    correctformat.append(templatesheet.cell(row=row, column=col).value)
                correctformat.pop(0)
                while correctformat[-1] is None:
                    del correctformat[-1]
                return correctformat
            
            
def renamereformat(wsdf, correctformat, renamedict, newcollist, mergecoldict):
    # renaming columns #
    for cname in renamedict:
        if cname in wsdf.columns and renamedict[cname] not in wsdf.columns:
            wsdf.rename(columns={cname:renamedict[cname]}, inplace=True)
    
    # creating new columns #
    for col in newcollist:
        if col not in wsdf.columns:
            wsdf[col] = ""
    
    # creating merge columns #
    for cname in mergecoldict: 
        wsdf[cname] = ""
        for mergecol in mergecoldict:
            for index in range(len(mergecoldict[mergecol])):
                wsdf[cname] += wsdf[mergecoldict[mergecol][index+1]].astype(str)
                if (index + 1) == (len(mergecoldict[mergecol]) - 1):
                    break
                else:
                    wsdf[cname] += mergecoldict[mergecol][0][0]
    
    # Removing blank rows #
    wsdf = wsdf.dropna(how='all')
    # creating df with corrected format #
    for colname in correctformat:
        if colname not in wsdf.columns:
            wsdf[colname] = ""
    wsdf = wsdf[correctformat].copy()
    # Removing identical rows #
    wsdf = wsdf.drop_duplicates()
    return wsdf

## wsdf compiler ##
def wsdfcompiler(dftobecompiled, wsdf):
    dftobecompiled = pd.concat([dftobecompiled, wsdf])
    return dftobecompiled

## ReFormat, Compiles, Index Match and Exports the final 'sheet' to xlsx format ## 
def rfcoimexsheet(ctemplatepath, sheetdict, formattype, cwstype, renamedict, newcollist, mergecoldict, iminstructdict):
    templatesheet = gettemplatesheet(ctemplatepath, formattype)
    correctformat = getcorrectformat(templatesheet, cwstype, formattype)
    dftobecompiled = pd.DataFrame(columns = [correctformat]) # dataframe to be compiled to the correct format #
    
    print('==== For type name: ',cwstype,' ====') ########################## console.log ####################
    for xlsxfile in glob.glob("./To be compiled/*.xlsx"):
        filename = xlsxfile.replace('./To be compiled\\','')
        if cwstype in sheetdict[filename]: 
            print(filename) ########################## console.log ####################
            wsdf = getrawdf(xlsxfile, cwstype)
            wsdf = renamereformat(wsdf, correctformat, renamedict, newcollist, mergecoldict)
            if dftobecompiled.empty == True:
                dftobecompiled = wsdf
            else:
                dftobecompiled = wsdfcompiler(dftobecompiled, wsdf)
    
    # Index Matching #
    iminstructdictc = checkindexmatch(dftobecompiled, iminstructdict)
    dftobecompiled = im.indexmatcher(dftobecompiled, iminstructdictc)
    
    return dftobecompiled

def finalsheetscompiler(ctemplatepath, formattype, cwstypelist, renamedict, newcollist, mergecoldict, compilermode, iminstructdict):
    sheetdict = asdl.getsheetdict()
    finalsheetsdict = dict()
#     if compilermode != 'polymerise':
    for cwstype in cwstypelist:
        finalsheetsdict[cwstype] = rfcoimexsheet(ctemplatepath, sheetdict, formattype, cwstype, renamedict, newcollist, mergecoldict, iminstructdict)
    return finalsheetsdict